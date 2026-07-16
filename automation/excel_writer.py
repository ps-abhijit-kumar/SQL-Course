from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font
class ExcelWriter:
    def __init__(self):
        self.workbook = Workbook()
        # Remove default sheet
        self.workbook.remove(
            self.workbook.active
        )
    def add_sheet(self, sheet_name, columns, rows):
        sheet = self.workbook.create_sheet(sheet_name)
        # Write headers
        for column_number, column_name in enumerate(columns, start=1):
            cell = sheet.cell(
                row=1,
                column=column_number
            )
            cell.value = column_name
            cell.font = Font(bold=True)
        # Write rows
        for row_number, row in enumerate(rows, start=2):
            for column_number, value in enumerate(row, start=1):
                sheet.cell(
                    row=row_number,
                    column=column_number
                ).value = value
        # Auto-size columns
        for column in sheet.columns:
            length = max(
                len(str(cell.value))
                if cell.value is not None else 0
                for cell in column
            )
            sheet.column_dimensions[
                column[0].column_letter
            ].width = length + 3
    def save(self):
        output_folder = (
            Path(__file__).parent /
            "output"
        )
        output_folder.mkdir(
            exist_ok=True
        )
        output_file = (
            output_folder /
            "sql_exercise_outputs.xlsx"
        )
        self.workbook.save(output_file)
        return output_file